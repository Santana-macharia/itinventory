# views.py
import csv
from io import TextIOWrapper
from django.shortcuts import render, redirect
from .forms import ImportForm
from .models import Import
import os
from django.conf import settings
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from reportlab.pdfgen import canvas
from django.http import HttpResponse
import openpyxl
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO  # Import BytesIO for handling the PDF file

def upload_csv(request):
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            # Check if 'file' is present in request.FILES
            if 'file' in request.FILES:
                handle_uploaded_file(request.FILES['file'])
                form.save()  # Save the form to create the Model instance
                return redirect('display_csv')
    else:
        form = ImportForm()
    return render(request, 'import/uploadcsv.html', {'form': form})

def handle_uploaded_file(file):
    decoded_file = TextIOWrapper(file.file, encoding='utf-8')
    reader = csv.reader(decoded_file)
    for row in reader:
        # Save the CSV data to the database or process as needed
        print(row)

    # Create the directory if it doesn't exist
    upload_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
    os.makedirs(upload_dir, exist_ok=True)

    # Save the file to a specific path within MEDIA_ROOT
    file_path = os.path.join(upload_dir, file.name)
    with open(file_path, 'wb') as destination:
        for chunk in file.chunks():
            destination.write(chunk)

def display_csv(request):
    # Retrieve data from the database (ImportModel)
    data = Import.objects.all()

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        data = data.filter(
            Q(centre__icontains=search_query) |
            Q(department__icontains=search_query) |
            Q(hardware__icontains=search_query) |
            # Q(device_name__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            # Q(assigned_by__icontains=search_query) |
            Q(assignee_first_name__icontains=search_query) |
            Q(assignee_last_name__icontains=search_query) | 
            Q(assignee_email_address__icontains=search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(date__icontains=search_query) 
        )

    # Set the number of items per page (100 in this case)
    items_per_page = 100
    paginator = Paginator(data, items_per_page)

    # Get the current page number from the request's GET parameters
    page_number = request.GET.get('page', 1)

    try:
        # Try to convert the page number to an integer
        page_number = int(page_number)
    except ValueError:
        # If the conversion fails, default to the first page
        page_number = 1

    try:
        # Get the data for the specified page
        data_on_page = paginator.page(page_number)
    except PageNotAnInteger:
        # If the page number is not an integer, default to the first page
        data_on_page = paginator.page(1)
    except EmptyPage:
        # If the page is out of range (e.g., greater than the number of pages), display the last page
        data_on_page = paginator.page(paginator.num_pages)

    # Generate a report based on the search query
    report_data = {
        'search_query': search_query,
        'total_records': data.count(),
    }

    return render(request, 'import/displaycsv.html', {
        'data': data_on_page,
        'paginator': paginator,
        'report_data': report_data,
    })

def displaycsv(request):
    # Your view logic for displaycsv.html goes here
    return render(request, 'devices/report/displaycsv.html')

# views.py 
def export_to_pdf(request):
    # Retrieve data from the database (ImportModel)
    data = Import.objects.all()

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        data = data.filter(
            Q(centre__icontains=search_query) |
            Q(department__icontains=search_query) |
            Q(hardware__icontains=search_query) |
            Q(system_model__icontains=search_query) |
            # Q(device_name__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            # Q(assigned_by__icontains=search_query) |
            Q(assignee_first_name__icontains=search_query) |
            Q(assignee_last_name__icontains=search_query) | 
            Q(assignee_email_address__icontains=search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(date__icontains=search_query) 
        )

    # Create a BytesIO buffer for the PDF file
    pdf_buffer = BytesIO()

    # Create a PDF document
    template_path = 'import/pdf.html'  # Replace 'your_app' with the actual app name
    template = get_template(template_path)
    html = template.render({'data': data})
    pisaStatus = pisa.CreatePDF(html, dest=pdf_buffer)

    # Check if PDF creation was successful
    if pisaStatus.err:
        return HttpResponse('Error creating PDF', content_type='text/plain')

    # Set up the response with PDF content
    pdf_buffer.seek(0)
    response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="exported_data.pdf"'

    # Close the PDF buffer
    pdf_buffer.close()

    return response


def export_to_excel(request):
    # Retrieve data from the database (ImportModel)
    data = Import.objects.all()

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        data = data.filter(
            Q(centre__icontains=search_query) |
            Q(department__icontains=search_query) |
            Q(hardware__icontains=search_query) |
            # Q(device_name__icontains=search_query) |
            Q(processor__icontains=search_query) |
            Q(ram_gb__icontains=search_query) |
            Q(hdd_gb__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            # Q(assigned_by__icontains=search_query) |
            Q(assignee_first_name__icontains=search_query) |
            Q(assignee_last_name__icontains=search_query) | 
            Q(assignee_email_address__icontains=search_query) |
            Q(device_condition__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(date__icontains=search_query) 
        )

    # Create a workbook and add a worksheet
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add headers to the worksheet
    headers = [
        'Centre', 'Department', 'Hardware', 'System Model', 'Device Name',
        'Processor', 'Ram (GB)', 'HDD (GB)', 'Serial Number', 'Assigned By',
        'Assignee First Name', 'Assignee Last Name', 'Assignee Email Address',
        'Device Condition', 'Status', 'Date'
    ]
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add data to the worksheet
    for row_num, item in enumerate(data, 2):
        worksheet.cell(row=row_num, column=1, value=item.centre)
        worksheet.cell(row=row_num, column=2, value=item.department)
        worksheet.cell(row=row_num, column=3, value=item.hardware)
        worksheet.cell(row=row_num, column=4, value=item.system_model)
        # worksheet.cell(row=row_num, column=5, value=item.device_name)
        worksheet.cell(row=row_num, column=6, value=item.processor)
        worksheet.cell(row=row_num, column=7, value=item.ram_gb)
        worksheet.cell(row=row_num, column=8, value=item.hdd_gb)
        worksheet.cell(row=row_num, column=9, value=item.serial_number)
        # worksheet.cell(row=row_num, column=10, value=item.assigned_by)
        worksheet.cell(row=row_num, column=11, value=item.assignee_first_name)
        worksheet.cell(row=row_num, column=12, value=item.assignee_last_name)
        worksheet.cell(row=row_num, column=13, value=item.assignee_email_address)
        worksheet.cell(row=row_num, column=14, value=item.device_condition)
        worksheet.cell(row=row_num, column=15, value=item.status)
        worksheet.cell(row=row_num, column=16, value=item.date)

    # Save the workbook to the response
    workbook.save(response)
    return response



# def upload_csv(request):
#     if request.method == 'POST':
#         form = ImportForm(request.POST, request.FILES)
#         if form.is_valid():
#             # Save the uploaded CSV file to the database
#             handle_uploaded_file(request.FILES['file'])
#             return redirect('display_csv')
#     else:
#         form = ImportForm()
#     return render(request, 'import/uploadcsv.html', {'form': form})

# def handle_uploaded_file(file):
#     # Process the uploaded CSV file and save data to the database
#     reader = csv.reader(file)
#     for row in reader:
#         _, created = Import.objects.get_or_create(centre=row[0], department=row[1], hardware=row[2], system_model=row[3],
#                                                    device_name =row[4],  processor=row[5], ram_gb =row[6], hdd_gb=row[7],
#                                                   serial_number=row[8], assigned_by=row[9], assignee_first_name=row[10],  assignee_last_name=row[11],
#                                                    assignee_email_address =row[12],  device_condition=row[13], status=row[14], date=row[15],)
#         # Add more fields as needed

# def display_csv(request):
#     # Retrieve data from the database and display it
#     data = Import.objects.all()
#     return render(request, 'import/displaycsv.html', {'data': data})

